from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.utils import timezone
from decimal import Decimal, InvalidOperation
import json
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from accounts.permissions import RolePermission
from .models import Product, Category, ProductUnit, Inventory, StockMovement, ProductSupplier
from suppliers.models import Supplier
from business.models import Branch
from core.pagination import StandardLimitOffsetPagination


class ProductBySkuView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"cashier", "salesperson", "supervisor", "admin"}

    def get(self, request, sku):
        try:
            product = Product.objects.get(sku=sku)
        except Product.DoesNotExist:
            return Response({"error": "Product not found"}, status=404)

        return Response(
            {
                "id": product.id,
                "name": product.name,
                "price": product.retail_price or product.selling_price,
            }
        )


def _build_product_queryset(request):
    branch_id = request.query_params.get("branch")
    query = (request.query_params.get("search") or "").strip()
    category_value = (request.query_params.get("category") or "").strip()
    products = Product.objects.select_related("category").prefetch_related("units")

    if branch_id:
        products = products.annotate(
            stock=Sum("inventory__quantity", filter=Q(inventory__branch_id=branch_id))
        )
    else:
        products = products.annotate(stock=Sum("inventory__quantity"))

    if query:
        products = products.filter(
            Q(name__icontains=query) | Q(sku__icontains=query)
        )

    if category_value:
        category_qs = Category.objects.filter(name__iexact=category_value)
        try:
            category_qs = category_qs | Category.objects.filter(id=category_value)
        except (ValueError, TypeError):
            pass
        category = category_qs.first()
        if category:
            products = products.filter(category=category)
        else:
            products = products.none()

    return products.order_by("name", "sku")


class CategoryListView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"cashier", "salesperson", "supervisor", "admin"}

    def get(self, request):
        query = (request.query_params.get("search") or "").strip()
        include_inactive = request.query_params.get("include_inactive") == "1"
        user = request.user
        role = (getattr(user, "role", "") or "").strip().lower()
        can_view_inactive = getattr(user, "is_superuser", False) or role in ("admin", "supervisor")

        categories = Category.objects.order_by("name")
        if not (include_inactive and can_view_inactive):
            categories = categories.filter(is_active=True)
        if query:
            categories = categories.filter(name__icontains=query)
        data = [
            {
                "id": str(c.id),
                "name": c.name,
                "is_active": c.is_active,
            }
            for c in categories
        ]
        return Response(data)


class CategoryCreateView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def post(self, request):
        data = request.data or {}
        errors = {}

        name = str(data.get("name") or "").strip()
        if not name:
            errors["name"] = "Category name is required."
        if name and Category.objects.filter(name__iexact=name).exists():
            errors["name"] = "Category already exists."

        is_active = _parse_bool(data.get("is_active"))

        if errors:
            return Response(errors, status=400)

        category = Category.objects.create(
            name=name,
            is_active=True if is_active is None else is_active,
        )
        return Response({"id": str(category.id)}, status=201)


class CategoryUpdateView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def put(self, request, category_id):
        category = get_object_or_404(Category, id=category_id)
        data = request.data or {}
        errors = {}

        name = str(data.get("name") or "").strip()
        if not name:
            errors["name"] = "Category name is required."
        elif Category.objects.filter(name__iexact=name).exclude(id=category.id).exists():
            errors["name"] = "Category already exists."

        is_active = _parse_bool(data.get("is_active"))

        if errors:
            return Response(errors, status=400)

        category.name = name
        if is_active is not None:
            category.is_active = is_active
        category.save()
        return Response({"id": str(category.id)}, status=200)


class ProductListView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"cashier", "salesperson", "supervisor", "admin"}

    def get(self, request):
        """List all active products with their category, prices, and current stock."""
        products = _build_product_queryset(request)

        paginator = StandardLimitOffsetPagination()
        page = paginator.paginate_queryset(products, request, view=self)
        page = page if page is not None else products

        data = []
        for product in page:
            total_stock = int(product.stock or 0)

            data.append({
                "id": str(product.id),
                "name": product.name,
                "sku": product.sku,
                "category_id": str(product.category_id) if product.category_id else None,
                "category": product.category.name if product.category else None,
                "cost_price": str(product.cost_price),
                "selling_price": str(product.retail_price or product.selling_price),
                "retail_price": str(product.retail_price or product.selling_price),
                "wholesale_price": str(product.wholesale_price) if product.wholesale_price is not None else None,
                "wholesale_threshold": product.wholesale_threshold,
                "is_active": product.is_active,
                "units": [
                    {
                        "id": str(u.id),
                        "unit_name": u.unit_name,
                        "unit_code": u.unit_code,
                        "conversion_to_base_unit": u.conversion_to_base_unit,
                        "is_base_unit": u.is_base_unit,
                        "retail_price": str(u.retail_price) if u.retail_price is not None else None,
                        "wholesale_price": str(u.wholesale_price) if u.wholesale_price is not None else None,
                        "wholesale_threshold": u.wholesale_threshold,
                    }
                    for u in product.units.filter(is_active=True)
                ],
                "stock": total_stock,
            })

        if page is not products:
            return paginator.get_paginated_response(data)
        return Response(data)


class ProductExportView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def get(self, request):
        products = _build_product_queryset(request)
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        headers = [
            "sku",
            "name",
            "category",
            "base_unit_name",
            "base_unit_code",
            "cost_price",
            "selling_price",
            "retail_price",
            "wholesale_price",
            "wholesale_threshold",
            "stock_quantity",
            "is_active",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for product in products:
            units = list(product.units.all())
            base_unit = next((unit for unit in units if unit.is_base_unit), None)
            ws.append([
                product.sku,
                product.name,
                product.category.name if product.category else "",
                base_unit.unit_name if base_unit else "",
                base_unit.unit_code if base_unit else "",
                product.cost_price,
                product.selling_price,
                product.retail_price if product.retail_price is not None else "",
                product.wholesale_price if product.wholesale_price is not None else "",
                product.wholesale_threshold if product.wholesale_threshold is not None else "",
                int(product.stock or 0),
                "Yes" if product.is_active else "No",
            ])

        for idx, header in enumerate(headers, start=1):
            width = max(14, len(header) + 2)
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"products_export_{timezone.localdate().isoformat()}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


def _get_branch_from_value(value):
    if value is None or str(value).strip() == "":
        return None
    branch_str = str(value).strip()
    if len(branch_str) >= 32:
        branch = Branch.objects.filter(id=branch_str).first()
        if branch:
            return branch
    return Branch.objects.filter(branch_name__iexact=branch_str).first()


def _ensure_base_unit(product, *, unit_name, retail_price, wholesale_price, wholesale_threshold):
    base_unit = product.units.filter(is_base_unit=True).first()
    if not base_unit:
        unit_code = (unit_name or "base").lower().replace(" ", "_")[:32]
        ProductUnit.objects.create(
            product=product,
            unit_name=unit_name or "Base Unit",
            unit_code=unit_code,
            conversion_to_base_unit=1,
            is_base_unit=True,
            retail_price=retail_price,
            wholesale_price=wholesale_price,
            wholesale_threshold=wholesale_threshold,
            is_active=True,
        )
        return
    if unit_name:
        base_unit.unit_name = unit_name
        base_unit.unit_code = unit_name.lower().replace(" ", "_")[:32]
    if retail_price is not None:
        base_unit.retail_price = retail_price
    if wholesale_price is not None:
        base_unit.wholesale_price = wholesale_price
    if wholesale_threshold is not None:
        base_unit.wholesale_threshold = wholesale_threshold
    base_unit.save()


class ProductCreateView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def post(self, request):
        data = request.data or {}
        errors = {}

        sku = str(data.get("sku") or "").strip()
        name = str(data.get("name") or "").strip()
        category_name = str(data.get("category") or "").strip()
        unit_name = str(data.get("unit") or "").strip()
        is_active = _parse_bool(data.get("is_active"))

        if not sku:
            errors["sku"] = "SKU is required."
        if not name:
            errors["name"] = "Name is required."
        if not category_name:
            errors["category"] = "Category is required."

        try:
            cost_price = _parse_decimal(data.get("cost_price"), field="cost_price")
        except ValueError as exc:
            cost_price = None
            errors["cost_price"] = str(exc)
        try:
            selling_price = _parse_decimal(data.get("selling_price"), field="selling_price")
        except ValueError as exc:
            selling_price = None
            errors["selling_price"] = str(exc)
        try:
            retail_price = _parse_decimal(data.get("retail_price"), field="retail_price")
        except ValueError as exc:
            retail_price = None
            errors["retail_price"] = str(exc)
        try:
            wholesale_price = _parse_decimal(data.get("wholesale_price"), field="wholesale_price")
        except ValueError as exc:
            wholesale_price = None
            errors["wholesale_price"] = str(exc)
        try:
            wholesale_threshold = _parse_int(data.get("wholesale_threshold"), field="wholesale_threshold")
        except ValueError as exc:
            wholesale_threshold = None
            errors["wholesale_threshold"] = str(exc)
        try:
            stock_quantity = _parse_int(data.get("stock_quantity"), field="stock_quantity")
        except ValueError as exc:
            stock_quantity = None
            errors["stock_quantity"] = str(exc)

        try:
            units_payload = _parse_units_payload(
                data.get("units"),
                retail_price=retail_price,
                selling_price=selling_price,
                wholesale_price=wholesale_price,
                wholesale_threshold=wholesale_threshold,
            )
        except ValueError as exc:
            units_payload = None
            errors["units"] = str(exc)

        branch = _get_branch_from_value(data.get("branch"))
        if stock_quantity is not None and not branch:
            errors["branch"] = "Branch is required when stock_quantity is provided."

        if cost_price is None:
            errors["cost_price"] = errors.get("cost_price") or "Cost price is required."
        if selling_price is None:
            errors["selling_price"] = errors.get("selling_price") or "Selling price is required."

        if sku and Product.objects.filter(sku=sku).exists():
            errors["sku"] = "SKU already exists."

        if errors:
            return Response(errors, status=400)

        with transaction.atomic():
            category, _ = Category.objects.get_or_create(name=category_name)
            product = Product.objects.create(
                sku=sku,
                name=name,
                category=category,
                cost_price=cost_price,
                selling_price=selling_price,
                retail_price=retail_price,
                wholesale_price=wholesale_price,
                wholesale_threshold=wholesale_threshold,
                is_active=True if is_active is None else is_active,
            )
            if units_payload:
                _sync_product_units(product, units_payload)
                base_unit = next((u for u in units_payload if u.get("is_base_unit")), None)
                if base_unit:
                    product.retail_price = base_unit.get("retail_price")
                    product.wholesale_price = base_unit.get("wholesale_price")
                    product.wholesale_threshold = base_unit.get("wholesale_threshold")
                    product.save(update_fields=["retail_price", "wholesale_price", "wholesale_threshold", "updated_at"])
            else:
                _ensure_base_unit(
                    product,
                    unit_name=unit_name,
                    retail_price=retail_price or selling_price,
                    wholesale_price=wholesale_price,
                    wholesale_threshold=wholesale_threshold,
                )
            if branch and stock_quantity is not None:
                Inventory.objects.update_or_create(
                    branch=branch,
                    product=product,
                    defaults={"quantity": stock_quantity},
                )

        return Response({"id": str(product.id)}, status=201)


class ProductUpdateView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def put(self, request, product_id):
        product = get_object_or_404(Product, id=product_id)
        data = request.data or {}
        errors = {}

        sku = str(data.get("sku") or product.sku).strip()
        name = str(data.get("name") or product.name).strip()
        category_name = str(data.get("category") or product.category.name).strip()
        unit_name = str(data.get("unit") or "").strip()
        is_active = _parse_bool(data.get("is_active"))

        if not sku:
            errors["sku"] = "SKU is required."
        if not name:
            errors["name"] = "Name is required."
        if not category_name:
            errors["category"] = "Category is required."

        try:
            cost_price = _parse_decimal(data.get("cost_price"), field="cost_price")
        except ValueError as exc:
            cost_price = None
            errors["cost_price"] = str(exc)
        try:
            selling_price = _parse_decimal(data.get("selling_price"), field="selling_price")
        except ValueError as exc:
            selling_price = None
            errors["selling_price"] = str(exc)
        try:
            retail_price = _parse_decimal(data.get("retail_price"), field="retail_price")
        except ValueError as exc:
            retail_price = None
            errors["retail_price"] = str(exc)
        try:
            wholesale_price = _parse_decimal(data.get("wholesale_price"), field="wholesale_price")
        except ValueError as exc:
            wholesale_price = None
            errors["wholesale_price"] = str(exc)
        try:
            wholesale_threshold = _parse_int(data.get("wholesale_threshold"), field="wholesale_threshold")
        except ValueError as exc:
            wholesale_threshold = None
            errors["wholesale_threshold"] = str(exc)
        try:
            stock_quantity = _parse_int(data.get("stock_quantity"), field="stock_quantity")
        except ValueError as exc:
            stock_quantity = None
            errors["stock_quantity"] = str(exc)

        try:
            units_payload = _parse_units_payload(
                data.get("units"),
                product=product,
                retail_price=retail_price,
                selling_price=selling_price,
                wholesale_price=wholesale_price,
                wholesale_threshold=wholesale_threshold,
            )
        except ValueError as exc:
            units_payload = None
            errors["units"] = str(exc)

        branch = _get_branch_from_value(data.get("branch"))
        if stock_quantity is not None and not branch:
            errors["branch"] = "Branch is required when stock_quantity is provided."

        if cost_price is None:
            errors["cost_price"] = errors.get("cost_price") or "Cost price is required."
        if selling_price is None:
            errors["selling_price"] = errors.get("selling_price") or "Selling price is required."

        if sku and Product.objects.filter(sku=sku).exclude(id=product.id).exists():
            errors["sku"] = "SKU already exists."

        if errors:
            return Response(errors, status=400)

        with transaction.atomic():
            category, _ = Category.objects.get_or_create(name=category_name)
            product.sku = sku
            product.name = name
            product.category = category
            product.cost_price = cost_price
            product.selling_price = selling_price
            product.retail_price = retail_price
            product.wholesale_price = wholesale_price
            product.wholesale_threshold = wholesale_threshold
            if is_active is not None:
                product.is_active = is_active
            product.save()

            if units_payload:
                _sync_product_units(product, units_payload)
                base_unit = next((u for u in units_payload if u.get("is_base_unit")), None)
                if base_unit:
                    product.retail_price = base_unit.get("retail_price")
                    product.wholesale_price = base_unit.get("wholesale_price")
                    product.wholesale_threshold = base_unit.get("wholesale_threshold")
                    product.save(update_fields=["retail_price", "wholesale_price", "wholesale_threshold", "updated_at"])
            else:
                _ensure_base_unit(
                    product,
                    unit_name=unit_name,
                    retail_price=retail_price or selling_price,
                    wholesale_price=wholesale_price,
                    wholesale_threshold=wholesale_threshold,
                )

            if branch and stock_quantity is not None:
                Inventory.objects.update_or_create(
                    branch=branch,
                    product=product,
                    defaults={"quantity": stock_quantity},
                )

        return Response({"id": str(product.id)}, status=200)


def _normalize_header(value):
    return (str(value).strip().lower() if value is not None else "")

def _normalize_category_name(value):
    if value is None:
        return ""
    name = str(value).strip()
    if not name:
        return ""
    name = re.sub(r"\s+", " ", name)
    return name.title()


def _parse_decimal(value, *, field):
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        raise ValueError(f"Invalid {field}")


def _parse_int(value, *, field):
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(Decimal(str(value).strip()))
    except (InvalidOperation, ValueError, TypeError):
        raise ValueError(f"Invalid {field}")


def _parse_bool(value):
    if value is None or str(value).strip() == "":
        return None
    val = str(value).strip().lower()
    if val in ("1", "true", "yes", "y", "on"):
        return True
    if val in ("0", "false", "no", "n", "off"):
        return False
    return None


def _normalize_unit_code(value):
    if value is None:
        return ""
    code = str(value).strip().lower()
    code = re.sub(r"\s+", "_", code)
    code = re.sub(r"[^a-z0-9_]", "", code)
    return code[:32]


def _coerce_units_payload(data):
    if data is None:
        return None
    if isinstance(data, list):
        return data
    if isinstance(data, str):
        try:
            parsed = json.loads(data)
        except json.JSONDecodeError as exc:
            raise ValueError("Units must be a JSON array.") from exc
        if not isinstance(parsed, list):
            raise ValueError("Units must be a list.")
        return parsed
    raise ValueError("Units must be a list.")


def _parse_units_payload(
    units_data,
    *,
    product=None,
    retail_price=None,
    selling_price=None,
    wholesale_price=None,
    wholesale_threshold=None,
):
    if units_data is None:
        return None
    units = _coerce_units_payload(units_data)
    cleaned = []
    base_count = 0
    codes = set()

    for idx, raw in enumerate(units, start=1):
        if not isinstance(raw, dict):
            raise ValueError(f"Unit entry {idx} must be an object.")

        unit_id = raw.get("id")
        if unit_id and product:
            if not product.units.filter(id=unit_id).exists():
                raise ValueError(f"Unit entry {idx} does not belong to this product.")

        unit_name = str(raw.get("unit_name") or "").strip()
        if not unit_name:
            raise ValueError(f"Unit entry {idx} is missing unit_name.")

        unit_code = str(raw.get("unit_code") or "").strip()
        if not unit_code:
            unit_code = _normalize_unit_code(unit_name) or "unit"
        unit_code = unit_code[:32]
        if unit_code in codes:
            raise ValueError(f"Duplicate unit code: {unit_code}.")
        codes.add(unit_code)

        try:
            conversion = _parse_int(raw.get("conversion_to_base_unit"), field="conversion_to_base_unit")
        except ValueError as exc:
            raise ValueError(f"Unit entry {idx}: {exc}") from exc
        if conversion is None or conversion <= 0:
            raise ValueError(f"Unit entry {idx}: conversion_to_base_unit is required and must be positive.")

        is_base = _parse_bool(raw.get("is_base_unit")) or False
        if is_base:
            base_count += 1
            if conversion != 1:
                raise ValueError("Base unit conversion must be 1.")

        retail = None
        wholesale = None
        threshold = None
        if raw.get("retail_price") not in (None, ""):
            retail = _parse_decimal(raw.get("retail_price"), field="retail_price")
        if raw.get("wholesale_price") not in (None, ""):
            wholesale = _parse_decimal(raw.get("wholesale_price"), field="wholesale_price")
        if raw.get("wholesale_threshold") not in (None, ""):
            threshold = _parse_int(raw.get("wholesale_threshold"), field="wholesale_threshold")

        is_active = _parse_bool(raw.get("is_active"))
        if is_active is None:
            is_active = True

        cleaned.append(
            {
                "id": unit_id,
                "unit_name": unit_name,
                "unit_code": unit_code,
                "conversion_to_base_unit": conversion,
                "is_base_unit": is_base,
                "retail_price": retail,
                "wholesale_price": wholesale,
                "wholesale_threshold": threshold,
                "is_active": is_active,
            }
        )

    if base_count != 1:
        raise ValueError("Exactly one base unit is required.")

    for unit in cleaned:
        if unit["is_base_unit"]:
            if unit["retail_price"] is None:
                unit["retail_price"] = retail_price if retail_price is not None else selling_price
            if unit["wholesale_price"] is None:
                unit["wholesale_price"] = wholesale_price
            if unit["wholesale_threshold"] is None:
                unit["wholesale_threshold"] = wholesale_threshold

    return cleaned


def _sync_product_units(product, units_payload):
    existing = {str(u.id): u for u in product.units.all()}
    keep_ids = set()

    for unit in units_payload:
        unit_id = str(unit.get("id") or "")
        if unit_id and unit_id in existing:
            obj = existing[unit_id]
        else:
            obj = ProductUnit(product=product)

        obj.unit_name = unit["unit_name"]
        obj.unit_code = unit["unit_code"]
        obj.conversion_to_base_unit = unit["conversion_to_base_unit"]
        obj.is_base_unit = unit["is_base_unit"]
        obj.retail_price = unit.get("retail_price")
        obj.wholesale_price = unit.get("wholesale_price")
        obj.wholesale_threshold = unit.get("wholesale_threshold")
        obj.is_active = unit.get("is_active", True)
        obj.save()
        keep_ids.add(str(obj.id))

    for obj in product.units.exclude(id__in=keep_ids):
        if obj.is_active:
            obj.is_active = False
            obj.save(update_fields=["is_active", "updated_at"])


class ProductImportTemplateView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def get(self, request):
        headers = [
            "sku",
            "name",
            "category",
            "cost_price",
            "selling_price",
            "retail_price",
            "wholesale_price",
            "wholesale_threshold",
            "unit",
            "stock_quantity",
            "branch",
            "is_active",
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(headers)
        for idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = max(14, len(header) + 2)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="product-import-template.xlsx"'
        wb.save(response)
        return response


class ProductImportView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def post(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "No file uploaded."}, status=400)
        if not upload.name.lower().endswith(".xlsx"):
            return Response({"detail": "Only .xlsx files are supported."}, status=400)

        try:
            wb = load_workbook(upload, data_only=True)
        except Exception:
            return Response({"detail": "Invalid Excel file."}, status=400)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Excel sheet is empty."}, status=400)

        headers = [_normalize_header(h) for h in rows[0]]
        header_map = {name: idx for idx, name in enumerate(headers) if name}

        required_fields = ["sku", "name", "category", "cost_price", "selling_price"]
        missing_headers = [h for h in required_fields if h not in header_map]
        if missing_headers:
            return Response(
                {"detail": f"Missing required headers: {', '.join(missing_headers)}"},
                status=400,
            )

        created_count = 0
        updated_count = 0
        failed_count = 0
        errors = []

        for row_index, row in enumerate(rows[1:], start=2):
            if not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue

            row_errors = []

            def add_error(field, message):
                row_errors.append({"row": row_index, "field": field, "error": message})

            def cell_value(field):
                idx = header_map.get(field)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            sku = str(cell_value("sku") or "").strip()
            name = str(cell_value("name") or "").strip()
            raw_category = cell_value("category")
            category_name = _normalize_category_name(raw_category)

            if not sku:
                add_error("sku", "Missing SKU")
            if not name:
                add_error("name", "Missing name")
            if not category_name:
                add_error("category", "Missing category")

            try:
                cost_price = _parse_decimal(cell_value("cost_price"), field="cost_price")
            except ValueError as exc:
                cost_price = None
                add_error("cost_price", str(exc))

            try:
                selling_price = _parse_decimal(cell_value("selling_price"), field="selling_price")
            except ValueError as exc:
                selling_price = None
                add_error("selling_price", str(exc))

            try:
                retail_price = _parse_decimal(cell_value("retail_price"), field="retail_price")
            except ValueError as exc:
                retail_price = None
                add_error("retail_price", str(exc))

            try:
                wholesale_price = _parse_decimal(cell_value("wholesale_price"), field="wholesale_price")
            except ValueError as exc:
                wholesale_price = None
                add_error("wholesale_price", str(exc))

            try:
                wholesale_threshold = _parse_int(cell_value("wholesale_threshold"), field="wholesale_threshold")
            except ValueError as exc:
                wholesale_threshold = None
                add_error("wholesale_threshold", str(exc))

            try:
                stock_quantity = _parse_int(cell_value("stock_quantity"), field="stock_quantity")
            except ValueError as exc:
                stock_quantity = None
                add_error("stock_quantity", str(exc))

            unit_name = str(cell_value("unit") or "").strip()
            is_active = _parse_bool(cell_value("is_active"))

            branch_value = cell_value("branch")
            branch = None
            if branch_value is not None and str(branch_value).strip() != "":
                branch_str = str(branch_value).strip()
                if len(branch_str) >= 32:
                    branch = Branch.objects.filter(id=branch_str).first()
                if branch is None:
                    branch = Branch.objects.filter(branch_name__iexact=branch_str).first()
                if branch is None:
                    add_error("branch", "Branch not found")
            elif stock_quantity is not None:
                add_error("branch", "Branch is required when stock_quantity is provided")

            if cost_price is None:
                add_error("cost_price", "Missing cost_price")
            if selling_price is None:
                add_error("selling_price", "Missing selling_price")

            if row_errors:
                errors.extend(row_errors)
                failed_count += 1
                continue

            with transaction.atomic():
                category = Category.objects.filter(name__iexact=category_name).first()
                if not category:
                    category = Category.objects.create(name=category_name)
                product, created = Product.objects.get_or_create(
                    sku=sku,
                    defaults={
                        "name": name,
                        "category": category,
                        "cost_price": cost_price,
                        "selling_price": selling_price,
                        "retail_price": retail_price,
                        "wholesale_price": wholesale_price,
                        "wholesale_threshold": wholesale_threshold,
                        "is_active": True if is_active is None else is_active,
                    },
                )

                if created:
                    created_count += 1
                else:
                    product.name = name
                    product.category = category
                    if cost_price is not None:
                        product.cost_price = cost_price
                    if selling_price is not None:
                        product.selling_price = selling_price
                    if retail_price is not None:
                        product.retail_price = retail_price
                    if wholesale_price is not None:
                        product.wholesale_price = wholesale_price
                    if wholesale_threshold is not None:
                        product.wholesale_threshold = wholesale_threshold
                    if is_active is not None:
                        product.is_active = is_active
                    product.save()
                    updated_count += 1

                base_unit = product.units.filter(is_base_unit=True).first()
                if not base_unit:
                    unit_code = unit_name.lower().replace(" ", "_")[:32] if unit_name else "base"
                    unit_label = unit_name or "Base Unit"
                    ProductUnit.objects.create(
                        product=product,
                        unit_name=unit_label,
                        unit_code=unit_code,
                        conversion_to_base_unit=1,
                        is_base_unit=True,
                        retail_price=retail_price or selling_price,
                        wholesale_price=wholesale_price,
                        wholesale_threshold=wholesale_threshold,
                        is_active=True,
                    )
                else:
                    if unit_name:
                        base_unit.unit_name = unit_name
                        base_unit.unit_code = unit_name.lower().replace(" ", "_")[:32]
                    if retail_price is not None:
                        base_unit.retail_price = retail_price
                    if wholesale_price is not None:
                        base_unit.wholesale_price = wholesale_price
                    if wholesale_threshold is not None:
                        base_unit.wholesale_threshold = wholesale_threshold
                    base_unit.save()

                if branch and stock_quantity is not None:
                    inventory, inv_created = Inventory.objects.get_or_create(
                        branch=branch,
                        product=product,
                        defaults={"quantity": stock_quantity},
                    )
                    if not inv_created:
                        inventory.quantity = stock_quantity
                        inventory.save()

        return Response(
            {
                "created_count": created_count,
                "updated_count": updated_count,
                "failed_count": failed_count,
                "errors": errors,
            }
        )


class InventoryStockLookupView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def get(self, request):
        product_id = request.query_params.get("product") or request.query_params.get("product_id")
        branch_id = request.query_params.get("branch") or request.query_params.get("branch_id")
        if not product_id or not branch_id:
            return Response({"detail": "Product and branch are required."}, status=400)

        inventory = Inventory.objects.filter(product_id=product_id, branch_id=branch_id).first()
        qty = inventory.quantity if inventory else 0
        return Response({"quantity": qty})


class InventoryAdjustmentListView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def get(self, request):
        branch_id = request.query_params.get("branch")
        product_id = request.query_params.get("product")

        qs = StockMovement.objects.filter(movement_type="adjustment").select_related(
            "product",
            "branch",
            "created_by",
        )
        if branch_id:
            qs = qs.filter(branch_id=branch_id)
        if product_id:
            qs = qs.filter(product_id=product_id)

        qs = qs.order_by("-created_at")

        paginator = StandardLimitOffsetPagination()
        page = paginator.paginate_queryset(qs, request, view=self)
        page = page if page is not None else qs

        data = [
            {
                "id": str(entry.id),
                "created_at": entry.created_at,
                "product_id": str(entry.product_id) if entry.product_id else None,
                "product_name": entry.product.name if entry.product else None,
                "product_sku": entry.product.sku if entry.product else None,
                "branch_id": str(entry.branch_id) if entry.branch_id else None,
                "branch_name": entry.branch.branch_name if entry.branch else None,
                "adjustment_type": "increase" if entry.quantity_change >= 0 else "decrease",
                "quantity": abs(entry.quantity_change),
                "previous_quantity": entry.previous_quantity,
                "new_quantity": entry.new_quantity,
                "reason": entry.reference or "",
                "note": entry.notes or "",
                "created_by_name": f"{entry.created_by.first_name} {entry.created_by.last_name}".strip()
                if entry.created_by
                else None,
            }
            for entry in page
        ]

        if page is not qs:
            return paginator.get_paginated_response(data)
        return Response(data)


class InventoryAdjustmentCreateView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def post(self, request):
        data = request.data or {}
        errors = {}

        product_id = data.get("product_id") or data.get("product")
        branch_id = data.get("branch_id") or data.get("branch")
        adj_type = (data.get("adjustment_type") or "").strip().lower()
        reason = str(data.get("reason") or "").strip()
        note = str(data.get("note") or "").strip()

        if not product_id:
            errors["product"] = "Product is required."
        if not branch_id:
            errors["branch"] = "Branch is required."
        if adj_type not in ("increase", "decrease"):
            errors["adjustment_type"] = "Adjustment type must be increase or decrease."
        if not reason:
            errors["reason"] = "Reason is required."

        try:
            quantity = _parse_int(data.get("quantity"), field="quantity")
        except ValueError as exc:
            quantity = None
            errors["quantity"] = str(exc)
        if quantity is None or quantity <= 0:
            errors["quantity"] = "Quantity must be greater than zero."

        product = Product.objects.filter(id=product_id).first() if product_id else None
        if product_id and not product:
            errors["product"] = "Product not found."
        branch = Branch.objects.filter(id=branch_id).first() if branch_id else None
        if branch_id and not branch:
            errors["branch"] = "Branch not found."

        if errors:
            return Response(errors, status=400)

        with transaction.atomic():
            inventory, _ = Inventory.objects.get_or_create(
                product=product,
                branch=branch,
                defaults={"quantity": 0},
            )
            previous_qty = inventory.quantity
            delta = quantity if adj_type == "increase" else -quantity
            new_qty = previous_qty + delta
            if new_qty < 0:
                return Response({"detail": "Insufficient stock for this decrease."}, status=400)

            inventory.quantity = new_qty
            inventory.save()

            movement = StockMovement.objects.create(
                inventory=inventory,
                product=product,
                branch=branch,
                movement_type="adjustment",
                quantity_change=delta,
                previous_quantity=previous_qty,
                new_quantity=new_qty,
                reference=reason[:100],
                notes=note or "",
                is_active=True,
            )

        return Response(
            {
                "id": str(movement.id),
                "product_id": str(product.id),
                "branch_id": str(branch.id),
                "previous_quantity": previous_qty,
                "new_quantity": new_qty,
            },
            status=201,
        )


class ProductSupplierListView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def get(self, request, product_id):
        product = get_object_or_404(Product, id=product_id)
        links = ProductSupplier.objects.filter(product=product).select_related("supplier").order_by("-is_primary", "supplier__name")
        data = [
            {
                "id": str(link.id),
                "product_id": str(product.id),
                "supplier": {
                    "id": str(link.supplier_id),
                    "name": link.supplier.name,
                    "phone": link.supplier.phone,
                    "email": link.supplier.email,
                    "contact_person": link.supplier.contact_person,
                    "is_active": link.supplier.is_active,
                },
                "supplier_sku": link.supplier_sku,
                "supplier_price": str(link.supplier_price) if link.supplier_price is not None else None,
                "is_primary": link.is_primary,
                "notes": link.notes,
            }
            for link in links
        ]
        return Response(data)


class ProductSupplierLinkView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def post(self, request, product_id):
        product = get_object_or_404(Product, id=product_id)
        data = request.data or {}
        errors = {}

        supplier_id = data.get("supplier_id") or data.get("supplier")
        supplier = Supplier.all_objects.filter(id=supplier_id).first() if supplier_id else None
        if not supplier:
            errors["supplier"] = "Supplier not found."

        supplier_sku = str(data.get("supplier_sku") or "").strip()
        notes = str(data.get("notes") or "").strip()

        try:
            supplier_price = _parse_decimal(data.get("supplier_price"), field="supplier_price")
        except ValueError as exc:
            errors["supplier_price"] = str(exc)
            supplier_price = None

        is_primary = _parse_bool(data.get("is_primary")) or False

        if errors:
            return Response(errors, status=400)

        if ProductSupplier.objects.filter(product=product, supplier=supplier).exists():
            return Response({"detail": "Supplier already linked to this product."}, status=400)

        if is_primary:
            ProductSupplier.objects.filter(product=product, is_primary=True).update(is_primary=False)

        link = ProductSupplier.objects.create(
            product=product,
            supplier=supplier,
            supplier_sku=supplier_sku,
            supplier_price=supplier_price,
            is_primary=is_primary,
            notes=notes,
        )
        return Response({"id": str(link.id)}, status=201)


class ProductSupplierUpdateView(APIView):
    permission_classes = [IsAuthenticated, RolePermission]
    allowed_roles = {"supervisor", "admin"}

    def put(self, request, product_id, link_id):
        product = get_object_or_404(Product, id=product_id)
        link = get_object_or_404(ProductSupplier, id=link_id, product=product)
        data = request.data or {}
        errors = {}

        supplier_sku = str(data.get("supplier_sku") or link.supplier_sku or "").strip()
        notes = str(data.get("notes") or link.notes or "").strip()

        try:
            supplier_price = _parse_decimal(data.get("supplier_price"), field="supplier_price")
        except ValueError as exc:
            errors["supplier_price"] = str(exc)
            supplier_price = None

        is_primary = _parse_bool(data.get("is_primary"))
        if is_primary is None:
            is_primary = link.is_primary

        if errors:
            return Response(errors, status=400)

        if is_primary and not link.is_primary:
            ProductSupplier.objects.filter(product=product, is_primary=True).exclude(id=link.id).update(is_primary=False)

        link.supplier_sku = supplier_sku
        link.notes = notes
        link.supplier_price = supplier_price
        link.is_primary = is_primary
        link.save()
        return Response({"id": str(link.id)}, status=200)

    def delete(self, request, product_id, link_id):
        product = get_object_or_404(Product, id=product_id)
        link = get_object_or_404(ProductSupplier, id=link_id, product=product)
        link.delete()
        return Response(status=204)
